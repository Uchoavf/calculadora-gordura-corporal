import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BRL_FORMAT = 'R$ #,##0.00'
DATE_FORMAT = 'DD/MM/YYYY'

MESES     = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
MESES_ABR = ["Jan","Fev","Mar","Abr","Mai","Jun",
             "Jul","Ago","Set","Out","Nov","Dez"]

CAT_REC  = ["Salário", "Freelance / Extra", "Investimentos", "Outros (Receita)"]
CAT_DESP = ["Moradia + Alimentação", "Transporte + Saúde", "Dívidas / Parcelas", "Outros (Despesa)"]

# Fixed row positions in each monthly sheet
REC_START      = 5
REC_END        = 8
TOTAL_REC_ROW  = 9
DESP_START     = 13
DESP_END       = 16
TOTAL_DESP_ROW = 17
SALDO_ROW      = 21

# Colors
CV_ESC = "1A5C38"; CV_MED = "2E7D32"; CV_CLR = "D5E8D4"
CR_ESC = "8B1A1A"; CR_MED = "C62828"; CR_CLR = "FFCDD2"
CA_ESC = "1B3A6B"; CA_MED = "1565C0"; CA_CLR = "E3F2FD"
CG_ESC = "37474F"; CG_CLR = "F5F5F5"; CBRC   = "FFFFFF"

def bold(sz=11, cor="000000"):
    return Font(bold=True, size=sz, color=cor, name="Calibri")

def normal(sz=11, cor="000000"):
    return Font(size=sz, color=cor, name="Calibri")

def bg(cor):
    return PatternFill(fill_type="solid", fgColor=cor)

_s = Side(style='thin')
THIN = Border(left=_s, right=_s, top=_s, bottom=_s)

def ctr(wrap=False):
    return Alignment(horizontal='center', vertical='center', wrap_text=wrap)

def rgt():
    return Alignment(horizontal='right', vertical='center')

def lft():
    return Alignment(horizontal='left', vertical='center')


def monthly_sheet(wb, mes, abr):
    ws = wb.create_sheet(title=abr)

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 3
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 34
    ws.column_dimensions['H'].width = 26
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 16

    # ── Título esquerdo ────────────────────────────────────────────────
    ws.merge_cells('A1:D1')
    c = ws['A1']; c.value = f"ORÇAMENTO — {mes.upper()} 2026"
    c.font = bold(15, CBRC); c.fill = bg(CA_ESC)
    c.alignment = ctr(); ws.row_dimensions[1].height = 32

    ws.row_dimensions[2].height = 6

    # ── Seção RECEITAS ─────────────────────────────────────────────────
    ws.merge_cells('A3:D3')
    c = ws['A3']; c.value = "▶  RECEITAS"
    c.font = bold(12, CBRC); c.fill = bg(CV_ESC)
    c.alignment = ctr(); ws.row_dimensions[3].height = 24

    for col, hdr in zip(['A','B','C','D'],
                        ['Categoria','Orçado (R$)','Realizado (R$)','Diferença (R$)']):
        c = ws[f'{col}4']
        c.value = hdr; c.font = bold(10, CBRC)
        c.fill = bg(CV_MED); c.alignment = ctr(True); c.border = THIN
    ws.row_dimensions[4].height = 28

    for i, cat in enumerate(CAT_REC):
        r = REC_START + i
        clr = CV_CLR if i % 2 == 0 else "EBF5EB"
        ws[f'A{r}'].value = cat; ws[f'A{r}'].font = normal()
        ws[f'A{r}'].fill = bg(clr); ws[f'A{r}'].border = THIN
        ws[f'A{r}'].alignment = lft(); ws.row_dimensions[r].height = 22
        for col in ['B','C']:
            c = ws[f'{col}{r}']; c.fill = bg(clr)
            c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()
        c = ws[f'D{r}']; c.value = f'=C{r}-B{r}'
        c.fill = bg(clr); c.border = THIN
        c.number_format = BRL_FORMAT; c.alignment = rgt()

    ws[f'A{TOTAL_REC_ROW}'].value = "TOTAL RECEITAS"
    ws[f'A{TOTAL_REC_ROW}'].font = bold(11, CBRC)
    ws[f'A{TOTAL_REC_ROW}'].fill = bg(CV_ESC)
    ws[f'A{TOTAL_REC_ROW}'].border = THIN
    ws[f'A{TOTAL_REC_ROW}'].alignment = lft()
    ws.row_dimensions[TOTAL_REC_ROW].height = 22
    for col in ['B','C','D']:
        c = ws[f'{col}{TOTAL_REC_ROW}']
        c.value = f'=SUM({col}{REC_START}:{col}{REC_END})'
        c.font = bold(11, CBRC); c.fill = bg(CV_ESC)
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()

    ws.row_dimensions[10].height = 6

    # ── Seção DESPESAS ─────────────────────────────────────────────────
    ws.merge_cells('A11:D11')
    c = ws['A11']; c.value = "▶  DESPESAS"
    c.font = bold(12, CBRC); c.fill = bg(CR_ESC)
    c.alignment = ctr(); ws.row_dimensions[11].height = 24

    for col, hdr in zip(['A','B','C','D'],
                        ['Categoria','Orçado (R$)','Realizado (R$)','Diferença (R$)']):
        c = ws[f'{col}12']
        c.value = hdr; c.font = bold(10, CBRC)
        c.fill = bg(CR_MED); c.alignment = ctr(True); c.border = THIN
    ws.row_dimensions[12].height = 28

    for i, cat in enumerate(CAT_DESP):
        r = DESP_START + i
        clr = CR_CLR if i % 2 == 0 else "FFE0E0"
        ws[f'A{r}'].value = cat; ws[f'A{r}'].font = normal()
        ws[f'A{r}'].fill = bg(clr); ws[f'A{r}'].border = THIN
        ws[f'A{r}'].alignment = lft(); ws.row_dimensions[r].height = 22
        for col in ['B','C']:
            c = ws[f'{col}{r}']; c.fill = bg(clr)
            c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()
        c = ws[f'D{r}']; c.value = f'=B{r}-C{r}'
        c.fill = bg(clr); c.border = THIN
        c.number_format = BRL_FORMAT; c.alignment = rgt()

    ws[f'A{TOTAL_DESP_ROW}'].value = "TOTAL DESPESAS"
    ws[f'A{TOTAL_DESP_ROW}'].font = bold(11, CBRC)
    ws[f'A{TOTAL_DESP_ROW}'].fill = bg(CR_ESC)
    ws[f'A{TOTAL_DESP_ROW}'].border = THIN
    ws[f'A{TOTAL_DESP_ROW}'].alignment = lft()
    ws.row_dimensions[TOTAL_DESP_ROW].height = 22
    for col in ['B','C','D']:
        c = ws[f'{col}{TOTAL_DESP_ROW}']
        c.value = f'=SUM({col}{DESP_START}:{col}{DESP_END})'
        c.font = bold(11, CBRC); c.fill = bg(CR_ESC)
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()

    ws.row_dimensions[18].height = 6

    # ── Seção SALDO ────────────────────────────────────────────────────
    ws.merge_cells('A19:D19')
    c = ws['A19']; c.value = "▶  SALDO DO MÊS"
    c.font = bold(12, CBRC); c.fill = bg(CA_ESC)
    c.alignment = ctr(); ws.row_dimensions[19].height = 24

    for col, hdr in zip(['A','B','C','D'],
                        ['','Orçado (R$)','Realizado (R$)','Diferença (R$)']):
        c = ws[f'{col}20']
        c.value = hdr; c.font = bold(10, CBRC)
        c.fill = bg(CA_MED); c.alignment = ctr(); c.border = THIN
    ws.row_dimensions[20].height = 22

    ws[f'A{SALDO_ROW}'].value = "Receitas − Despesas"
    ws[f'A{SALDO_ROW}'].font = bold(11)
    ws[f'A{SALDO_ROW}'].fill = bg(CA_CLR)
    ws[f'A{SALDO_ROW}'].border = THIN
    ws[f'A{SALDO_ROW}'].alignment = lft()
    ws.row_dimensions[SALDO_ROW].height = 24
    for col in ['B','C','D']:
        c = ws[f'{col}{SALDO_ROW}']
        c.value = f'={col}{TOTAL_REC_ROW}-{col}{TOTAL_DESP_ROW}'
        c.font = bold(11); c.fill = bg(CA_CLR)
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()

    # ── Lançamentos (lado direito) ─────────────────────────────────────
    ws.merge_cells('F1:J1')
    c = ws['F1']; c.value = f"LANÇAMENTOS — {mes.upper()} 2026"
    c.font = bold(13, CBRC); c.fill = bg(CG_ESC); c.alignment = ctr()

    ws.merge_cells('F2:J2')
    c = ws['F2']
    c.value = "Preencha as colunas B e C (Orçado e Realizado) nos cards à esquerda"
    c.font = normal(9, "555555"); c.alignment = ctr()
    ws.row_dimensions[2].height = 18

    for col, hdr in zip(['F','G','H','I','J'],
                        ['Data','Descrição','Categoria','Tipo','Valor (R$)']):
        c = ws[f'{col}3']
        c.value = hdr; c.font = bold(10, CBRC)
        c.fill = bg(CG_ESC); c.alignment = ctr(); c.border = THIN

    for r in range(4, 54):
        clr = CG_CLR if r % 2 == 0 else CBRC
        for col in ['F','G','H','I','J']:
            c = ws[f'{col}{r}']; c.fill = bg(clr); c.border = THIN
        ws[f'F{r}'].number_format = DATE_FORMAT
        ws[f'F{r}'].alignment = ctr()
        ws[f'J{r}'].number_format = BRL_FORMAT
        ws[f'J{r}'].alignment = rgt()
        ws.row_dimensions[r].height = 19

    dv = DataValidation(type="list", formula1='"Receita,Despesa"',
                        allow_blank=True, showDropDown=False,
                        showErrorMessage=True,
                        errorTitle="Tipo inválido",
                        error="Selecione Receita ou Despesa")
    ws.add_data_validation(dv)
    dv.sqref = "I4:I53"

    cats_str = '","'.join(CAT_REC + CAT_DESP)
    dv_cat = DataValidation(type="list",
                            formula1=f'Categorias!$A$2:$A$9',
                            allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_cat)
    dv_cat.sqref = "H4:H53"

    ws.freeze_panes = 'F2'
    return ws


def summary_sheet(wb):
    ws = wb.create_sheet(title="Resumo Anual", index=0)

    ws.column_dimensions['A'].width = 30
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.merge_cells('A1:N1')
    c = ws['A1']; c.value = "RESUMO ANUAL 2026 — ORÇAMENTO FINANCEIRO"
    c.font = bold(16, CBRC); c.fill = bg(CA_ESC)
    c.alignment = ctr(); ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:N2')
    c = ws['A2']; c.value = "Valores Realizados por Categoria • Coluna C de cada mês"
    c.font = normal(10, CBRC); c.fill = bg(CA_MED)
    c.alignment = ctr(); ws.row_dimensions[2].height = 18

    # Cabeçalho de colunas
    ws['A3'].value = "Categoria"; ws['A3'].font = bold(10, CBRC)
    ws['A3'].fill = bg(CG_ESC); ws['A3'].border = THIN; ws['A3'].alignment = ctr()
    ws.row_dimensions[3].height = 22
    for j, abr in enumerate(MESES_ABR):
        col = get_column_letter(j + 2)
        c = ws[f'{col}3']; c.value = abr; c.font = bold(10, CBRC)
        c.fill = bg(CG_ESC); c.border = THIN; c.alignment = ctr()
    ws['N3'].value = "TOTAL ANUAL"; ws['N3'].font = bold(10, CBRC)
    ws['N3'].fill = bg(CA_ESC); ws['N3'].border = THIN; ws['N3'].alignment = ctr()

    # Receitas
    ws.merge_cells('A4:N4')
    c = ws['A4']; c.value = "RECEITAS"; c.font = bold(12, CBRC)
    c.fill = bg(CV_ESC); c.alignment = ctr(); ws.row_dimensions[4].height = 22

    for i, cat in enumerate(CAT_REC):
        r = 5 + i; clr = CV_CLR if i % 2 == 0 else "EBF5EB"
        ws[f'A{r}'].value = cat; ws[f'A{r}'].font = normal()
        ws[f'A{r}'].fill = bg(clr); ws[f'A{r}'].border = THIN
        ws[f'A{r}'].alignment = lft(); ws.row_dimensions[r].height = 22
        for j, abr in enumerate(MESES_ABR):
            col = get_column_letter(j + 2)
            c = ws[f'{col}{r}']
            c.value = f"='{abr}'!C{REC_START + i}"
            c.fill = bg(clr); c.border = THIN
            c.number_format = BRL_FORMAT; c.alignment = rgt()
        c = ws[f'N{r}']; c.value = f'=SUM(B{r}:M{r})'
        c.font = bold(11); c.fill = bg("C8E6C9")
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()

    ws['A9'].value = "TOTAL RECEITAS"; ws['A9'].font = bold(11, CBRC)
    ws['A9'].fill = bg(CV_ESC); ws['A9'].border = THIN
    ws['A9'].alignment = lft(); ws.row_dimensions[9].height = 22
    for j, abr in enumerate(MESES_ABR):
        col = get_column_letter(j + 2)
        c = ws[f'{col}9']; c.value = f"='{abr}'!C{TOTAL_REC_ROW}"
        c.font = bold(11, CBRC); c.fill = bg(CV_ESC)
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()
    ws['N9'].value = '=SUM(B9:M9)'; ws['N9'].font = bold(11, CBRC)
    ws['N9'].fill = bg(CV_ESC); ws['N9'].border = THIN
    ws['N9'].number_format = BRL_FORMAT; ws['N9'].alignment = rgt()

    ws.row_dimensions[10].height = 6

    # Despesas
    ws.merge_cells('A11:N11')
    c = ws['A11']; c.value = "DESPESAS"; c.font = bold(12, CBRC)
    c.fill = bg(CR_ESC); c.alignment = ctr(); ws.row_dimensions[11].height = 22

    for i, cat in enumerate(CAT_DESP):
        r = 12 + i; clr = CR_CLR if i % 2 == 0 else "FFE0E0"
        ws[f'A{r}'].value = cat; ws[f'A{r}'].font = normal()
        ws[f'A{r}'].fill = bg(clr); ws[f'A{r}'].border = THIN
        ws[f'A{r}'].alignment = lft(); ws.row_dimensions[r].height = 22
        for j, abr in enumerate(MESES_ABR):
            col = get_column_letter(j + 2)
            c = ws[f'{col}{r}']; c.value = f"='{abr}'!C{DESP_START + i}"
            c.fill = bg(clr); c.border = THIN
            c.number_format = BRL_FORMAT; c.alignment = rgt()
        c = ws[f'N{r}']; c.value = f'=SUM(B{r}:M{r})'
        c.font = bold(11); c.fill = bg("FFCDD2")
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()

    ws['A16'].value = "TOTAL DESPESAS"; ws['A16'].font = bold(11, CBRC)
    ws['A16'].fill = bg(CR_ESC); ws['A16'].border = THIN
    ws['A16'].alignment = lft(); ws.row_dimensions[16].height = 22
    for j, abr in enumerate(MESES_ABR):
        col = get_column_letter(j + 2)
        c = ws[f'{col}16']; c.value = f"='{abr}'!C{TOTAL_DESP_ROW}"
        c.font = bold(11, CBRC); c.fill = bg(CR_ESC)
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()
    ws['N16'].value = '=SUM(B16:M16)'; ws['N16'].font = bold(11, CBRC)
    ws['N16'].fill = bg(CR_ESC); ws['N16'].border = THIN
    ws['N16'].number_format = BRL_FORMAT; ws['N16'].alignment = rgt()

    ws.row_dimensions[17].height = 6

    # Saldo
    ws.merge_cells('A18:N18')
    c = ws['A18']; c.value = "SALDO MENSAL  (Receitas − Despesas)"
    c.font = bold(12, CBRC); c.fill = bg(CA_ESC)
    c.alignment = ctr(); ws.row_dimensions[18].height = 22

    ws['A19'].value = "Saldo Realizado"; ws['A19'].font = bold(11)
    ws['A19'].fill = bg(CA_CLR); ws['A19'].border = THIN
    ws['A19'].alignment = lft(); ws.row_dimensions[19].height = 24
    for j, abr in enumerate(MESES_ABR):
        col = get_column_letter(j + 2)
        c = ws[f'{col}19']; c.value = f"='{abr}'!C{SALDO_ROW}"
        c.font = bold(11); c.fill = bg(CA_CLR)
        c.border = THIN; c.number_format = BRL_FORMAT; c.alignment = rgt()
    ws['N19'].value = '=SUM(B19:M19)'; ws['N19'].font = bold(11)
    ws['N19'].fill = bg(CA_CLR); ws['N19'].border = THIN
    ws['N19'].number_format = BRL_FORMAT; ws['N19'].alignment = rgt()

    # Saldo acumulado
    ws['A20'].value = "Saldo Acumulado"; ws['A20'].font = normal(10, "555555")
    ws['A20'].fill = bg("E8EAF6"); ws['A20'].border = THIN
    ws['A20'].alignment = lft(); ws.row_dimensions[20].height = 20
    ws['B20'].value = '=B19'
    ws['B20'].fill = bg("E8EAF6"); ws['B20'].border = THIN
    ws['B20'].number_format = BRL_FORMAT; ws['B20'].alignment = rgt()
    for j in range(1, 12):
        col = get_column_letter(j + 2)
        prev = get_column_letter(j + 1)
        c = ws[f'{col}20']; c.value = f'={prev}20+{col}19'
        c.fill = bg("E8EAF6"); c.border = THIN
        c.number_format = BRL_FORMAT; c.alignment = rgt()
    ws['N20'].value = ""; ws['N20'].fill = bg("E8EAF6"); ws['N20'].border = THIN

    ws.freeze_panes = 'B4'
    return ws


def categories_sheet(wb):
    ws = wb.create_sheet(title="Categorias")
    ws.sheet_state = 'hidden'
    ws['A1'].value = "Categorias"
    ws['A1'].font = bold()
    all_cats = CAT_REC + CAT_DESP
    for i, cat in enumerate(all_cats):
        ws[f'A{i+2}'].value = cat
    ws.column_dimensions['A'].width = 30
    return ws


def instructions_sheet(wb):
    ws = wb.create_sheet(title="Como Usar", index=0)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60

    ws.merge_cells('A1:B1')
    c = ws['A1']; c.value = "📋  COMO USAR ESTE ORÇAMENTO"
    c.font = bold(16, CBRC); c.fill = bg(CA_ESC)
    c.alignment = ctr(); ws.row_dimensions[1].height = 38

    rows = [
        ("", ""),
        ("ESTRUTURA DA PLANILHA", "header"),
        ("Resumo Anual", "• Visão geral de todos os meses em uma única tela."),
        ("Abas Jan – Dez",
         "• Uma aba por mês com orçamento e lançamentos lado a lado."),
        ("", ""),
        ("COMO PREENCHER CADA MÊS", "header"),
        ("Passo 1 – Orçado",
         "• Coluna B de cada categoria: quanto você PLANEJA gastar/receber."),
        ("Passo 2 – Realizado",
         "• Coluna C de cada categoria: quanto você EFETIVAMENTE gastou/recebeu."),
        ("Diferença",
         "• Calculada automaticamente (Coluna D)."),
        ("Saldo do Mês",
         "• Calculado automaticamente: Total Receitas − Total Despesas."),
        ("", ""),
        ("LANÇAMENTOS DETALHADOS (lado direito)", "header"),
        ("Data",         "• Formato DD/MM/AAAA."),
        ("Descrição",    "• Ex: 'Conta de luz', 'Salário março'."),
        ("Categoria",    "• Selecione da lista suspensa."),
        ("Tipo",         "• Selecione: Receita ou Despesa."),
        ("Valor (R$)",   "• Sempre positivo."),
        ("", ""),
        ("DICAS", "header"),
        ("Saldo verde",  "• Receitas > Despesas — ótimo!"),
        ("Saldo vermelho","• Despesas > Receitas — revise o orçamento."),
        ("Acumulado",    "• Linha 'Saldo Acumulado' no Resumo mostra o crescimento ao longo do ano."),
    ]

    for i, (col_a, col_b) in enumerate(rows):
        r = i + 2
        if col_b == "header":
            ws.merge_cells(f'A{r}:B{r}')
            c = ws[f'A{r}']; c.value = col_a
            c.font = bold(12, CBRC); c.fill = bg(CG_ESC)
            c.alignment = Alignment(horizontal='left', vertical='center',
                                    indent=1)
            ws.row_dimensions[r].height = 24
        elif col_a == "":
            ws.row_dimensions[r].height = 8
        else:
            ws[f'A{r}'].value = col_a
            ws[f'A{r}'].font = bold(10, CA_ESC)
            ws[f'A{r}'].alignment = Alignment(horizontal='right',
                                               vertical='center')
            ws[f'B{r}'].value = col_b
            ws[f'B{r}'].font = normal(10)
            ws[f'B{r}'].alignment = Alignment(horizontal='left',
                                               vertical='center', indent=1,
                                               wrap_text=True)
            clr = CG_CLR if r % 2 == 0 else CBRC
            ws[f'A{r}'].fill = bg(clr); ws[f'B{r}'].fill = bg(clr)
            ws.row_dimensions[r].height = 20
    return ws


# ── MAIN ──────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

categories_sheet(wb)
for mes, abr in zip(MESES, MESES_ABR):
    monthly_sheet(wb, mes, abr)
summary_sheet(wb)
instructions_sheet(wb)

out = "/home/user/calculadora-gordura-corporal/orcamento-financeiro-2026.xlsx"
wb.save(out)
print(f"Planilha gerada: {out}")
