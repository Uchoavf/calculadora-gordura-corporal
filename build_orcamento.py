import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BRL = '"R$"#,##0.00'
DATE = 'DD/MM/YYYY'

MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
MESES_ABR = ["Jan","Fev","Mar","Abr","Mai","Jun",
             "Jul","Ago","Set","Out","Nov","Dez"]

# ── CORES ──────────────────────────────────────────────────────────────────
CV_E="1A5C38"; CV_M="388E3C"; CV_C="E8F5E9"   # verde (receitas)
CO_E="BF6900"; CO_M="E65100"; CO_C="FFF3E0"   # laranja (descontos)
CR_E="7B1818"; CR_M="C62828"; CR_C="FFEBEE"   # vermelho (fixas/cheque)
CM_E="4A148C"; CM_M="7B1FA2"; CM_C="F3E5F5"   # roxo (cartões)
CA_E="1B3A6B"; CA_M="1565C0"; CA_C="E3F2FD"   # azul (títulos/resultado)
CG_E="37474F"; CG_C="F5F5F5"; CW="FFFFFF"     # cinza / branco

def bold(sz=10, c="000000"): return Font(bold=True,size=sz,color=c,name="Calibri")
def norm(sz=10, c="000000"): return Font(size=sz,color=c,name="Calibri")
def bg(c): return PatternFill(fill_type="solid",fgColor=c)
_t=Side(style='thin'); _m=Side(style='medium')
THIN=Border(left=_t,right=_t,top=_t,bottom=_t)
def ctr(w=False): return Alignment(horizontal='center',vertical='center',wrap_text=w)
def rgt(): return Alignment(horizontal='right',vertical='center')
def lft(i=0): return Alignment(horizontal='left',vertical='center',indent=i)

# ── DADOS PRÉ-PREENCHIDOS ───────────────────────────────────────────────────
RECEITAS  = ["SALÁRIO","Kalena OF/NET","ALUGUÉL","Wagner","Neto","UAB","",""]
DESCONTOS = ["Consignado V.Nbk","Caixa Habitação","Consignado BB","Emp.Sal.BB",
             "AUX.TRANSP","CONTRI.SOCIAL","IRRF","AUX.PRÉ.ESC","Dízimo","","",""]
FIXAS     = ["HAPVIDA/SALMINHO","HAPVIDA/SABRINA","HAP EWE E EFRA","CELPA ENERGIA",
             "SEATELECOM","TX.ÁLAMO","CONTA TIM","CONTA CLARO","MAXÍDOMINI",
             "COSANPA/ÁGUA","ESCOLA SALMOM","CORECON","Boticário","Rodobens","",""]
CARTOES   = ["BB","NU","Santander SX","SELITE","Caixa Gold","Cartão MP",
             "PicPay Card","Itaú Card","CAIXA ELO","Bradesco","RecargaPay",
             "BipaCard","InfinitCard","Pagbank","TonCard",""]
ASSINAT   = [("NU","GDRIVE SABRI",49.90,"DIA 04","CRÉDITO (NU)"),
             ("NU","CONTA TIM",42.00,"DIA 15","PG MANUAL"),
             ("SANTANDER SX","PLANO CLARO",73.73,"DIA 09","PG MANUAL"),
             ("NU","YOUTUBE ST",53.90,"DIA 25","CRÉDITO (NU)"),
             ("NU","ICLOUD",5.90,"DIA 01","CRÉDITO (NU)"),
             ("","Serasa",23.90,"DIA 17",""),
             ("","Kindle",24.00,"",""),
             ("NU","GPT",130.00,"DIA 02","CRÉDITO (NU)"),
             ("NU","Copilot",109.90,"DIA 04","CRÉDITO (NU)"),
             ("NU","WORD",0,"Anual","CRÉDITO (NU)"),
             ("NU","GDRIVE EWE",0,"Ago/26","CRÉDITO (NU)"),
             ("","PDFVIEW",0,"","CRÉDITO (NU)")]
CHQ_BANCOS = ["BB","Santander","Caixa","Itaú","Bradesco"]
CONTAS    = ["Saldo CC BB","Santander","Caixa","Bipa","Nubank","Itaú","MP",
             "KuCoin","Ton","Picpay","Recargapay","Infinitepay","PagBank",
             "C6 BANK","INTER","Bradesco","BTG"]
NOTAS     = [("Consignado Nubank 1","1.30% am","16.74% aa","73x"),
             ("Consignado Nubank 2","1.57% am","21.85% aa","96x — saldo R$776,82"),
             ("Fin. Habitação CAIXA","7% nominal","",""),
             ("Empréstimo Sabrina","","","saldo R$2.503,46"),
             ("Empréstimo Josy","","",""),
             ("","","",""),
             ("","","","")]

# ── ÂNCORAS DE LINHA (fixas para fórmulas cruzadas) ────────────────────────
DS=4; DE=19; TR=20          # data start/end, total row
AH=22; ACS=23; ADS=24; ADE=35; AT=36   # assinaturas
CH=38; CCS=39; CDS=40; CDE=44; CT=45   # cheque especial
RH=47                        # resultado header
RR=48; RD=49; RL=50; RF=51; RC=52; RJ=53; RCH=54; RA=55; RS=56; RND=57
CNH=59; CNCS=60; CNDS=61; CNDE=77; CNT=78  # contas


def cell(ws, ref, val=None, fnt=None, fil=None, aln=None, brd=THIN, fmt=None, h=None):
    c = ws[ref]
    if val is not None: c.value = val
    if fnt: c.font = fnt
    if fil: c.fill = fil
    if aln: c.alignment = aln
    if brd is not None: c.border = brd
    if fmt: c.number_format = fmt
    return c


def hdr_row(ws, row, sections, h=22):
    """sections: list of (col_start, col_end, text, bg_color)"""
    ws.row_dimensions[row].height = h
    for cs, ce, txt, col in sections:
        if cs != ce:
            ws.merge_cells(f'{cs}{row}:{ce}{row}')
        c = ws[f'{cs}{row}']
        c.value = txt; c.font = bold(11, CW)
        c.fill = bg(col); c.alignment = ctr(); c.border = THIN


def build_monthly(wb, mes, abr):
    ws = wb.create_sheet(title=abr)

    # larguras de coluna
    cw = {'A':24,'B':14,'C':13,'D':12,
          'E':24,'F':14,'G':13,'H':13,
          'I':24,'J':12,'K':9,'L':12,'M':13,'N':12,
          'O':16,'P':14,'Q':14,'R':12,'S':9,'T':12,'U':13,'V':16}
    for col, w in cw.items():
        ws.column_dimensions[col].width = w

    # ── LINHA 1: título principal ──────────────────────────────────────────
    ws.merge_cells('A1:V1')
    c = ws['A1']
    c.value = f"ORÇAMENTO  {mes.upper()}  2026"
    c.font = bold(14, CW); c.fill = bg(CA_E); c.alignment = ctr()
    ws.row_dimensions[1].height = 30

    # ── LINHA 2: cabeçalhos de seção ──────────────────────────────────────
    ws.row_dimensions[2].height = 24
    hdr_row(ws, 2, [
        ('A','D','▶  FONTES DE RECEITA',     CV_E),
        ('E','H','▶  DESCONTOS EM FOLHA',    CO_E),
        ('I','N','▶  DESPESAS FIXAS',        CR_E),
        ('O','V','▶  DESPESAS COM CARTÕES',  CM_E),
    ])

    # ── LINHA 3: sub-cabeçalhos ────────────────────────────────────────────
    ws.row_dimensions[3].height = 26
    subcols = [
        ('A',CV_M,'ORIGEM'),('B',CV_M,'VALOR'),('C',CV_M,'SITUAÇÃO'),('D',CV_M,'DATA PG'),
        ('E',CO_M,'ORIGEM'),('F',CO_M,'VALOR'),('G',CO_M,'INFO'),  ('H',CO_M,'OBS'),
        ('I',CR_M,'ORIGEM'),('J',CR_M,'VALOR'),('K',CR_M,'JUROS'), ('L',CR_M,'CET'),
        ('M',CR_M,'SITUAÇÃO'),('N',CR_M,'DATA PG'),
        ('O',CM_M,'CARTÃO'),('P',CM_M,'FATURA'),('Q',CM_M,'PAGO'),
        ('R',CM_M,'SALDO¹'),('S',CM_M,'JUROS'),('T',CM_M,'CET'),
        ('U',CM_M,'SITUAÇÃO'),('V',CM_M,'DATA PG'),
    ]
    for col, cor, txt in subcols:
        c = ws[f'{col}3']
        c.value=txt; c.font=bold(9,CW); c.fill=bg(cor); c.alignment=ctr(True); c.border=THIN

    # ── LINHAS 4-19: dados ────────────────────────────────────────────────
    for i in range(DE - DS + 1):
        r = DS + i
        ws.row_dimensions[r].height = 20
        alt = i % 2

        def row_cells(cols_clrs):
            for col, clr in cols_clrs:
                ws[f'{col}{r}'].fill = bg(clr); ws[f'{col}{r}'].border = THIN

        # Receitas
        rec_clr = CV_C if not alt else "F0FBF0"
        row_cells([('A',rec_clr),('B',rec_clr),('C',rec_clr),('D',rec_clr)])
        ws[f'A{r}'].value = RECEITAS[i] if i < len(RECEITAS) else ""
        ws[f'A{r}'].font = norm(); ws[f'A{r}'].alignment = lft(1)
        ws[f'B{r}'].number_format = BRL; ws[f'B{r}'].alignment = rgt()
        ws[f'D{r}'].number_format = DATE; ws[f'D{r}'].alignment = ctr()

        # Descontos
        dsc_clr = CO_C if not alt else "FFF8F0"
        row_cells([('E',dsc_clr),('F',dsc_clr),('G',dsc_clr),('H',dsc_clr)])
        ws[f'E{r}'].value = DESCONTOS[i] if i < len(DESCONTOS) else ""
        ws[f'E{r}'].font = norm(); ws[f'E{r}'].alignment = lft(1)
        ws[f'F{r}'].number_format = BRL; ws[f'F{r}'].alignment = rgt()

        # Fixas
        fix_clr = CR_C if not alt else "FFF5F5"
        row_cells([('I',fix_clr),('J',fix_clr),('K',fix_clr),('L',fix_clr),('M',fix_clr),('N',fix_clr)])
        ws[f'I{r}'].value = FIXAS[i] if i < len(FIXAS) else ""
        ws[f'I{r}'].font = norm(); ws[f'I{r}'].alignment = lft(1)
        ws[f'J{r}'].number_format = BRL; ws[f'J{r}'].alignment = rgt()
        ws[f'K{r}'].number_format = BRL; ws[f'K{r}'].alignment = rgt()
        ws[f'L{r}'].value = f'=IF(J{r}="","",J{r}+IF(K{r}="",0,K{r}))'
        ws[f'L{r}'].number_format = BRL; ws[f'L{r}'].alignment = rgt()
        ws[f'M{r}'].alignment = ctr()
        ws[f'N{r}'].number_format = DATE; ws[f'N{r}'].alignment = ctr()

        # Cartões — PAGO e SALDO calculados automaticamente
        car_clr = CM_C if not alt else "FAF0FF"
        row_cells([('O',car_clr),('P',car_clr),('Q',car_clr),('R',car_clr),
                   ('S',car_clr),('T',car_clr),('U',car_clr),('V',car_clr)])
        ws[f'O{r}'].value = CARTOES[i] if i < len(CARTOES) else ""
        ws[f'O{r}'].font = norm(); ws[f'O{r}'].alignment = lft(1)
        ws[f'P{r}'].number_format = BRL; ws[f'P{r}'].alignment = rgt()   # Fatura
        ws[f'Q{r}'].number_format = BRL; ws[f'Q{r}'].alignment = rgt()   # Pago
        ws[f'R{r}'].value = f'=IF(P{r}="","",P{r}-IF(Q{r}="",0,Q{r}))'  # Saldo auto
        ws[f'R{r}'].number_format = BRL; ws[f'R{r}'].alignment = rgt()
        ws[f'S{r}'].number_format = BRL; ws[f'S{r}'].alignment = rgt()   # Juros
        ws[f'T{r}'].value = f'=IF(Q{r}="","",Q{r}+IF(S{r}="",0,S{r}))'  # CET auto
        ws[f'T{r}'].number_format = BRL; ws[f'T{r}'].alignment = rgt()
        ws[f'U{r}'].alignment = ctr()
        ws[f'V{r}'].number_format = DATE; ws[f'V{r}'].alignment = ctr()

    # ── LINHA 20: TOTAIS (com fórmulas) ──────────────────────────────────
    ws.row_dimensions[TR].height = 22
    totais = [
        [('A',CV_E,'RECEITA TOTAL',lft(1)),('B',CV_E,f'=SUM(B{DS}:B{DE})',rgt()),
         ('C',CV_E,None,None),('D',CV_E,None,None)],
        [('E',CO_E,'TOTAL DESCONTOS',lft(1)),('F',CO_E,f'=SUM(F{DS}:F{DE})',rgt()),
         ('G',CO_E,None,None),('H',CO_E,None,None)],
        [('I',CR_E,'TOTAL FIXAS (CET)',lft(1)),('J',CR_E,f'=SUM(J{DS}:J{DE})',rgt()),
         ('K',CR_E,f'=SUM(K{DS}:K{DE})',rgt()),('L',CR_E,f'=SUM(L{DS}:L{DE})',rgt()),
         ('M',CR_E,None,None),('N',CR_E,None,None)],
        [('O',CM_E,'TOTAL CARTÕES',lft(1)),
         ('P',CM_E,f'=SUM(P{DS}:P{DE})',rgt()),  # total faturas
         ('Q',CM_E,f'=SUM(Q{DS}:Q{DE})',rgt()),  # total pago
         ('R',CM_E,f'=SUM(R{DS}:R{DE})',rgt()),  # total saldo (nova dívida)
         ('S',CM_E,f'=SUM(S{DS}:S{DE})',rgt()),  # total juros
         ('T',CM_E,f'=SUM(T{DS}:T{DE})',rgt()),  # total CET
         ('U',CM_E,None,None),('V',CM_E,None,None)],
    ]
    for grp in totais:
        for col, cor, val, aln in grp:
            c = ws[f'{col}{TR}']
            c.fill = bg(cor); c.border = THIN
            c.font = bold(10, CW)
            if val: c.value = val
            if aln: c.alignment = aln
            if col in ('B','F','J','K','L','P','Q','R','S','T'):
                c.number_format = BRL

    # nota de rodapé SALDO¹
    ws.merge_cells(f'O{TR+1}:V{TR+1}')
    ws[f'O{TR+1}'].value = "¹ SALDO = Fatura − Pago  (valor não quitado = nova dívida do mês)"
    ws[f'O{TR+1}'].font = Font(size=8, color="888888", italic=True)
    ws[f'O{TR+1}'].border = Border()

    # ── ASSINATURAS (A22:D36) + NOTAS EXPLICATIVAS (F22:I30) ─────────────
    ws.row_dimensions[AH].height = 22
    ws.merge_cells(f'A{AH}:D{AH}')
    ws[f'A{AH}'].value="ASSINATURAS ONLINE"; ws[f'A{AH}'].font=bold(11,CW)
    ws[f'A{AH}'].fill=bg(CG_E); ws[f'A{AH}'].alignment=ctr()

    ws.merge_cells(f'F{AH}:I{AH}')
    ws[f'F{AH}'].value="NOTAS / EMPRÉSTIMOS / CONSIGNADOS"
    ws[f'F{AH}'].font=bold(11,CW); ws[f'F{AH}'].fill=bg(CG_E)
    ws[f'F{AH}'].alignment=ctr()

    for col,hdr in zip(['A','B','C','D'],['CARTÃO','SERVIÇO','VALOR','DIA/SITUAÇÃO']):
        c=ws[f'{col}{ACS}']; c.value=hdr; c.font=bold(9,CW)
        c.fill=bg("546E7A"); c.alignment=ctr(True); c.border=THIN
    ws.row_dimensions[ACS].height = 20

    for col,hdr in zip(['F','G','H','I'],['DESCRIÇÃO','TAXA (am)','TAXA (aa)','OBS / SALDO']):
        c=ws[f'{col}{ACS}']; c.value=hdr; c.font=bold(9,CW)
        c.fill=bg("546E7A"); c.alignment=ctr(True); c.border=THIN

    for i,(cart,serv,val,dia,sit) in enumerate(ASSINAT):
        r=ADS+i; clr=CG_C if i%2==0 else CW
        ws.row_dimensions[r].height=18
        for col,v,fmt,aln in [('A',cart,None,ctr()),('B',serv,None,lft(1)),
                                ('C',val,BRL,rgt()),('D',f'{dia} — {sit}' if dia else sit,None,lft(1))]:
            c=ws[f'{col}{r}']; c.value=v; c.fill=bg(clr)
            c.border=THIN; c.font=norm()
            if fmt: c.number_format=fmt
            if aln: c.alignment=aln

    for i,(desc,am,aa,obs) in enumerate(NOTAS):
        r=ADS+i; clr=CG_C if i%2==0 else CW
        for col,v in zip(['F','G','H','I'],[desc,am,aa,obs]):
            c=ws[f'{col}{r}']; c.value=v; c.fill=bg(clr)
            c.border=THIN; c.font=norm()
            c.alignment=lft(1) if col in('F','I') else ctr()

    ws.row_dimensions[AT].height=20
    ws.merge_cells(f'A{AT}:B{AT}')
    ws[f'A{AT}'].value="TOTAL ASSINATURAS"; ws[f'A{AT}'].font=bold(10,CW)
    ws[f'A{AT}'].fill=bg(CG_E); ws[f'A{AT}'].border=THIN; ws[f'A{AT}'].alignment=lft(1)
    ws[f'C{AT}'].value=f'=SUM(C{ADS}:C{ADE})'
    ws[f'C{AT}'].font=bold(10,CW); ws[f'C{AT}'].fill=bg(CG_E)
    ws[f'C{AT}'].border=THIN; ws[f'C{AT}'].number_format=BRL; ws[f'C{AT}'].alignment=rgt()
    ws[f'D{AT}'].fill=bg(CG_E); ws[f'D{AT}'].border=THIN

    # ── CHEQUE ESPECIAL (A38:G45) ─────────────────────────────────────────
    ws.row_dimensions[CH].height=22
    ws.merge_cells(f'A{CH}:G{CH}')
    ws[f'A{CH}'].value="CHEQUE ESPECIAL / DÍVIDAS FINANCEIRAS"
    ws[f'A{CH}'].font=bold(11,CW); ws[f'A{CH}'].fill=bg(CR_E); ws[f'A{CH}'].alignment=ctr()

    for col,hdr in zip(['A','B','C','D','E','F','G'],
                       ['BANCO','LIMITE','USADO','JUROS MÊS','TAXA % aa','DIAS','SITUAÇÃO']):
        c=ws[f'{col}{CCS}']; c.value=hdr; c.font=bold(9,CW)
        c.fill=bg(CR_M); c.alignment=ctr(True); c.border=THIN
    ws.row_dimensions[CCS].height=22

    for i,banco in enumerate(CHQ_BANCOS):
        r=CDS+i; clr=CR_C if i%2==0 else "FFF5F5"
        ws.row_dimensions[r].height=19
        ws[f'A{r}'].value=banco
        for col in ['A','B','C','D','E','F','G']:
            c=ws[f'{col}{r}']; c.fill=bg(clr); c.border=THIN; c.font=norm()
        ws[f'A{r}'].alignment=lft(1)
        for col in ['B','C','D']:
            ws[f'{col}{r}'].number_format=BRL; ws[f'{col}{r}'].alignment=rgt()
        ws[f'E{r}'].alignment=ctr(); ws[f'F{r}'].alignment=ctr(); ws[f'G{r}'].alignment=ctr()

    ws.row_dimensions[CT].height=20
    ws.merge_cells(f'A{CT}:B{CT}')
    ws[f'A{CT}'].value="TOTAL CHEQUE ESPECIAL"; ws[f'A{CT}'].font=bold(10,CW)
    ws[f'A{CT}'].fill=bg(CR_E); ws[f'A{CT}'].border=THIN; ws[f'A{CT}'].alignment=lft(1)
    for col,formula in [('C',f'=SUM(C{CDS}:C{CDE})'),('D',f'=SUM(D{CDS}:D{CDE})')]:
        ws[f'{col}{CT}'].value=formula; ws[f'{col}{CT}'].font=bold(10,CW)
        ws[f'{col}{CT}'].fill=bg(CR_E); ws[f'{col}{CT}'].border=THIN
        ws[f'{col}{CT}'].number_format=BRL; ws[f'{col}{CT}'].alignment=rgt()
    for col in ['E','F','G']:
        ws[f'{col}{CT}'].fill=bg(CR_E); ws[f'{col}{CT}'].border=THIN

    # ── RESULTADO DO MÊS (A47:F57) ────────────────────────────────────────
    ws.row_dimensions[RH].height=28
    ws.merge_cells(f'A{RH}:F{RH}')
    ws[f'A{RH}'].value="  ══  RESULTADO DO MÊS  ══"
    ws[f'A{RH}'].font=bold(14,CW); ws[f'A{RH}'].fill=bg(CA_E); ws[f'A{RH}'].alignment=ctr()

    linhas=[
        (RR, "Receita Bruta",               f'=B{TR}',                           CV_C,  CV_E,  False),
        (RD, "(−) Descontos em Folha",       f'=F{TR}',                           CO_C,  CO_E,  False),
        (RL, "= Receita Líquida",            f'=F{RR}-F{RD}',                     CA_C,  CA_M,  True ),
        (RF, "(−) Despesas Fixas (CET)",     f'=L{TR}',                           CR_C,  CR_M,  False),
        (RC, "(−) Cartões — Valor Pago",     f'=Q{TR}',                           CM_C,  CM_M,  False),
        (RJ, "(−) Juros dos Cartões",        f'=S{TR}',                           CM_C,  CM_M,  False),
        (RCH,"(−) Juros Cheque Especial",    f'=D{CT}',                           CR_C,  CR_M,  False),
        (RA, "(−) Assinaturas Online",       f'=C{AT}',                           CG_C,  CG_E,  False),
        (RS, "= SALDO REAL DO MÊS",
             f'=F{RL}-F{RF}-F{RC}-F{RJ}-F{RCH}-F{RA}',                          CA_C,  CA_E,  True ),
        (RND,"⚠  Nova Dívida (cartões não pagos)", f'=R{TR}',                     "FFEBEE",CR_E,False),
    ]
    for row,label,formula,clr_r,clr_l,is_total in linhas:
        ws.row_dimensions[row].height = 24 if is_total else 19
        ws.merge_cells(f'A{row}:E{row}')
        c=ws[f'A{row}']; c.value=label
        c.font = bold(11,CW) if is_total else norm(10)
        c.fill = bg(clr_l if is_total else clr_r)
        c.border=THIN; c.alignment=lft(1)
        c=ws[f'F{row}']; c.value=formula
        c.font = bold(11,CW) if is_total else norm(10)
        c.fill = bg(clr_l if is_total else clr_r)
        c.border=THIN; c.number_format=BRL; c.alignment=rgt()

    # ── CONTAS E SALDOS (A59:B78) ─────────────────────────────────────────
    ws.row_dimensions[CNH].height=22
    ws.merge_cells(f'A{CNH}:C{CNH}')
    ws[f'A{CNH}'].value="CONTAS E SALDOS"
    ws[f'A{CNH}'].font=bold(11,CW); ws[f'A{CNH}'].fill=bg(CG_E); ws[f'A{CNH}'].alignment=ctr()

    for col,hdr in zip(['A','B','C'],['BANCO / CARTEIRA','SALDO (R$)','OBS']):
        c=ws[f'{col}{CNCS}']; c.value=hdr; c.font=bold(9,CW)
        c.fill=bg("546E7A"); c.alignment=ctr(True); c.border=THIN
    ws.row_dimensions[CNCS].height=18

    for i,conta in enumerate(CONTAS):
        r=CNDS+i; clr=CG_C if i%2==0 else CW
        ws.row_dimensions[r].height=18
        ws[f'A{r}'].value=conta
        for col in ['A','B','C']:
            ws[f'{col}{r}'].fill=bg(clr); ws[f'{col}{r}'].border=THIN; ws[f'{col}{r}'].font=norm()
        ws[f'A{r}'].alignment=lft(1)
        ws[f'B{r}'].number_format=BRL; ws[f'B{r}'].alignment=rgt()

    ws.row_dimensions[CNT].height=20
    ws[f'A{CNT}'].value="TOTAL EM CONTAS"; ws[f'A{CNT}'].font=bold(10,CW)
    ws[f'A{CNT}'].fill=bg(CG_E); ws[f'A{CNT}'].border=THIN; ws[f'A{CNT}'].alignment=lft(1)
    ws[f'B{CNT}'].value=f'=SUM(B{CNDS}:B{CNDE})'
    ws[f'B{CNT}'].font=bold(10,CW); ws[f'B{CNT}'].fill=bg(CG_E)
    ws[f'B{CNT}'].border=THIN; ws[f'B{CNT}'].number_format=BRL; ws[f'B{CNT}'].alignment=rgt()
    ws[f'C{CNT}'].fill=bg(CG_E); ws[f'C{CNT}'].border=THIN

    ws.freeze_panes='A4'
    return ws


def build_resumo(wb):
    ws = wb.create_sheet(title="Resumo 2026", index=0)

    ws.column_dimensions['A'].width = 28
    for i in range(2, 15):
        ws.column_dimensions[get_column_letter(i)].width = 13

    ws.merge_cells('A1:N1')
    ws['A1'].value = "RESUMO ANUAL 2026  —  ORÇAMENTO FINANCEIRO"
    ws['A1'].font = bold(15, CW); ws['A1'].fill = bg(CA_E)
    ws['A1'].alignment = ctr(); ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:N2')
    ws['A2'].value = ("Valores calculados automaticamente a partir das abas mensais  "
                      "•  SALDO¹ = valor não quitado nos cartões (nova dívida gerada)")
    ws['A2'].font = Font(size=9, color=CW, italic=True)
    ws['A2'].fill = bg(CA_M); ws['A2'].alignment = ctr()
    ws.row_dimensions[2].height = 18

    # cabeçalho de meses
    ws['A3'].value = ""; ws['A3'].fill = bg(CG_E); ws['A3'].border = THIN
    ws.row_dimensions[3].height = 22
    for j, abr in enumerate(MESES_ABR):
        col = get_column_letter(j + 2)
        c = ws[f'{col}3']; c.value = abr; c.font = bold(10, CW)
        c.fill = bg(CG_E); c.border = THIN; c.alignment = ctr()
    ws['N3'].value = "TOTAL ANO"; ws['N3'].font = bold(10, CW)
    ws['N3'].fill = bg(CA_E); ws['N3'].border = THIN; ws['N3'].alignment = ctr()

    # linhas do resumo — cada uma referencia célula específica de cada aba mensal
    metricas = [
        ("RECEITA BRUTA",          f'B{TR}',   CV_C, CV_E, False),
        ("(−) Descontos em Folha", f'F{TR}',   CO_C, CO_E, False),
        ("= Receita Líquida",      f'F{RL}',   CA_C, CA_M, True ),
        ("(−) Despesas Fixas CET", f'L{TR}',   CR_C, CR_M, False),
        ("(−) Cartões — Pago",     f'Q{TR}',   CM_C, CM_M, False),
        ("(−) Juros Cartões",      f'S{TR}',   CM_C, CM_M, False),
        ("(−) Juros Cheque Esp.",  f'D{CT}',   CR_C, CR_E, False),
        ("(−) Assinaturas",        f'C{AT}',   CG_C, CG_E, False),
        ("= SALDO REAL DO MÊS",   f'F{RS}',   CA_C, CA_E, True ),
        ("Nova Dívida (cartões)",  f'R{TR}',   "FFEBEE",CR_E,False),
        ("Total em Contas",        f'B{CNT}',  CG_C, CG_E, False),
    ]

    for i, (label, ref, clr_r, clr_l, is_tot) in enumerate(metricas):
        r = 4 + i
        ws.row_dimensions[r].height = 22 if is_tot else 19
        c = ws[f'A{r}']; c.value = label
        c.font = bold(10, CW) if is_tot else norm(10)
        c.fill = bg(clr_l if is_tot else clr_r); c.border = THIN; c.alignment = lft(1)

        for j, abr in enumerate(MESES_ABR):
            col = get_column_letter(j + 2)
            cell_ref = f"='{abr}'!{ref}"
            c = ws[f'{col}{r}']; c.value = cell_ref
            c.font = bold(10, CW) if is_tot else norm(10)
            c.fill = bg(clr_l if is_tot else clr_r); c.border = THIN
            c.number_format = BRL; c.alignment = rgt()

        # Total anual
        data_cols = f'B{r}:M{r}'
        c = ws[f'N{r}']; c.value = f'=SUM({data_cols})'
        c.font = bold(11, CW) if is_tot else bold(10, CW)
        c.fill = bg(clr_l); c.border = THIN
        c.number_format = BRL; c.alignment = rgt()

    ws.freeze_panes = 'B4'
    return ws


# ── MAIN ────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

for mes, abr in zip(MESES, MESES_ABR):
    build_monthly(wb, mes, abr)

build_resumo(wb)

out = "/home/user/calculadora-gordura-corporal/orcamento-2026-melhorado.xlsx"
wb.save(out)
print("Gerado:", out)
